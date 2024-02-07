"""
**Note:**   When the application is unexpectedly closed, the excel file may be corrupted!
            Please make sure to stop the application before closing.
            If you expect abrupt closures, please use version 1.

**Note:**   The starting directory is `data`.

**Note:**   If the file with the name `filepath` already exists, a new file is created.

**Note:**   Directories specified in the `filepath` (e.g. `dir1/dir2/test.xlsx`) are created automatically.

When the excel output module is added to the configuration,
the data changes are saved in the following way in the specified file:

| Timestamp            | Measurement   | Field key   | Field value   | Tag key    | Tag value            |
| :--------------------|:--------------|:------------| :-------------| :----------| :--------------------|
| 10-11-2020 18:21:26  | Rest_client   | level       | INFO          | host       | http://test.com      |

All tags are appended in new columns in the same row.
If one measurement has multiple fields, the fields are written separately in new rows.
"""
import os
import threading
from dataclasses import dataclass, field
import pathlib

# Internal imports.
import config
from modules.outputs.base.base import AbstractOutputModule, models


class OutputModule(AbstractOutputModule):
    """
    Class for the console output module.

    :param configuration: The configuration object of the module.
    """
    version: str = "1.0"
    """The version of the module."""
    public: bool = False
    """Is this module public?"""
    description: str = "This module stores the data in a specified excel file."
    """A short description."""
    author: str = "Colin Reiff"
    """The author name."""
    email: str = "colin.reiff@collectu.de"
    """The email address of the author."""
    deprecated: bool = True
    """Is this module deprecated."""
    third_party_requirements: list[str] = ["openpyxl==3.0.6"]
    """Define your requirements here."""
    can_be_buffer: bool = False
    """If True, the child has to implement 'store_buffer_data' and 'get_buffer_data'."""

    @dataclass
    class Configuration(models.OutputModule):
        """
        The configuration model of the module.
        """
        filepath: str = field(
            metadata=dict(description="The path and file name.",
                          required=False),
            default="excel_file/test.xls")
        include_tags: bool = field(
            metadata=dict(description="Shall tags be saved as well.",
                          required=False),
            default=True)

    def __init__(self, configuration: Configuration):
        # Calls the base init method.
        super().__init__(configuration=configuration)
        self.file = None
        """The path to the file (including the file name)."""
        self.workbook = None
        """The workbook."""
        self.sheet = None
        """The sheet."""
        self.row_index = 1
        """The current row index."""

    @classmethod
    def import_third_party_requirements(cls) -> bool:
        """
        Check if all third party requirements are successfully imported.
        Raises an ImportError if the import was not successful.

        :returns: True if the import was successful.
        """
        try:
            # Import third party packages here and mark them as global.
            global openpyxl
            import openpyxl
            return True
        except Exception:
            raise ImportError("Could not import required packages. Please install '{0}'."
                              .format(' '.join(map(str, cls.third_party_requirements))))

    def start(self) -> bool:
        """
        Just starts the thread for processing the queue.

        :returns: True if successfully connected, otherwise false.
        """
        try:
            # This is the file path including the file name.
            self.file = os.path.join('..', 'data', '{0}').format(self.configuration.filepath)
            # Create directory.
            pathlib.Path(os.path.join('..', 'data',
                                      str(pathlib.Path(self.configuration.filepath).parents[0]))).mkdir(
                parents=True,
                exist_ok=True)
            # Create new Workbook.
            self.workbook = openpyxl.Workbook()
            self.sheet = self.workbook.active
            self.sheet.title = self.configuration.id

            # Check if file already exists. If so, we add a number to the end.
            i = 1
            org_file_name = os.path.splitext(os.path.basename(self.file))[0]
            while os.path.isfile(self.file):
                temp_file_name = org_file_name + '_{0}'.format(str(i))
                self.file = os.path.join(os.path.dirname(self.file),
                                         temp_file_name + os.path.splitext(os.path.basename(self.file))[1])
                i += 1
            else:
                # Write the headings.
                self.sheet.cell(row=1, column=1, value="Timestamp")
                self.sheet.cell(row=1, column=2, value="Measurement")
                self.sheet.cell(row=1, column=3, value="Field key")
                self.sheet.cell(row=1, column=4, value="Field value")
                self.sheet.cell(row=1, column=5, value="Tag key")
                self.sheet.cell(row=1, column=6, value="Tag value")
                self.workbook.save(self.file)

            # Start the queue processing for storing incoming data.
            threading.Thread(target=self._process_queue,
                             daemon=False,
                             name="Queue_Worker_{0}".format(self.configuration.id)).start()

            self.logger.info("Successfully created file '{0}'.".format(self.file))
            return True

        except Exception as e:
            self.logger.critical("Could not create file '{0}'. {1}"
                                 .format(self.configuration.filepath, str(e)),
                                 exc_info=config.EXC_INFO)
            return False

    def _run(self, data: models.Data):
        """
        Method called when new data has to be processed.

        :param data: The data object to be processed.
        """
        try:
            # Make new row for each field.
            for field_key, field_value in data.fields.items():
                # Convert value if it is unknown (since module can not handle lists etc.).
                if not type(field_value) in [int, str, float, bool]:
                    field_value = str(field_value)
                self.row_index += 1
                self.sheet.cell(row=self.row_index, column=1, value=str(data.time))
                self.sheet.cell(row=self.row_index, column=2, value=data.measurement)
                self.sheet.cell(row=self.row_index, column=3, value=field_key)
                self.sheet.cell(row=self.row_index, column=4, value=field_value)
                if self.configuration.include_tags:
                    column_index = 0
                    # Add all tags at the end.
                    for tag_key, tag_value in data.tags.items():
                        if not type(tag_value) in [int, str, float, bool]:
                            tag_value = str(tag_value)
                        self.sheet.cell(row=self.row_index, column=5 + column_index, value=tag_key)
                        self.sheet.cell(row=self.row_index, column=6 + column_index, value=tag_value)
                        column_index += 2
            self.workbook.save(self.file)
        except Exception as e:
            self.logger.error("Something went wrong while trying to store data: {0}."
                              .format(str(e)),
                              exc_info=config.EXC_INFO)
            self._buffer(data=data, invalid=True)
